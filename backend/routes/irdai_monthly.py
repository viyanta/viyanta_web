from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy import text, bindparam
from sqlalchemy.orm import Session
from databases.database import get_db, DB_TYPE
import os
import re
from datetime import datetime
import calendar
from openpyxl import load_workbook
import tempfile
from services.irdai_excel_importer import import_irdai_excel

router = APIRouter()


DB_CONFIG = {
    "host": os.getenv("DB_HOST", "localhost"),
    "user": os.getenv("DB_USER", "root"),
    "password": os.getenv("DB_PASSWORD", ""),
    "database": os.getenv("DB_NAME", "viyanta_web"),
    "port": int(os.getenv("DB_PORT", 3306))
}
# ======================================================
# UPLOAD IRDAI MONTHLY EXCEL
# ======================================================


def extract_report_month_from_filename(filename: str) -> str:
    """
    Example:
    FYP Aug 2023.xlsx → 2023-08-31
    IRDAI December 2024.xlsx → 2024-12-31
    """
    match = re.search(
        r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+(\d{4})',
        filename,
        re.IGNORECASE
    )

    if not match:
        raise ValueError("Cannot extract report month from filename")

    month_str, year = match.groups()
    month = datetime.strptime(month_str[:3], "%b").month
    year = int(year)

    last_day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-{last_day:02d}"


@router.post("/upload-monthly-excel")
async def upload_irdai_monthly_excel(
    file: UploadFile = File(...)
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        # 1️⃣ AUTO-DETECT SHEET NAME
        wb = load_workbook(tmp_path, read_only=True)
        sheet_name = wb.sheetnames[0]
        wb.close()

        # 2️⃣ AUTO-DETECT REPORT MONTH
        report_month = extract_report_month_from_filename(file.filename)

        # 3️⃣ CALL EXISTING IMPORT LOGIC (UNCHANGED)
        result = import_irdai_excel(
            excel_path=tmp_path,
            sheet_name=sheet_name,
            report_month=report_month,
            db_config=DB_CONFIG,
        )

        return {
            **result,
            "sheet_name": sheet_name,
            "source_file": file.filename,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

# ======================================================
# 1️⃣ PERIOD TYPES
# ======================================================


@router.get("/period/types")
def get_period_types():
    return [
        {"label": "Monthly", "value": "MONTH"},
        {"label": "Quarterly", "value": "Q"},
        {"label": "Halfyearly", "value": "H"},
        {"label": "Annual", "value": "FY"},
    ]


# ======================================================
# 2️⃣ PERIOD OPTIONS
# ======================================================
@router.get("/period/options")
def get_period_options(
    type: str = Query(..., description="MONTH | Q | H | FY"),
    db: Session = Depends(get_db)
):
    if type == "MONTH":
        sql = text("""
            SELECT
              month_year AS label,
              MIN(report_month) AS start_date,
              MAX(report_month) AS end_date
            FROM irdai_monthly_data
            GROUP BY month_year
            ORDER BY MAX(report_month) DESC
        """)

    elif type == "Q":
        if DB_TYPE == "mysql":
            sql = text("""
                SELECT
                  CONCAT(
                    DATE_FORMAT(MIN(report_month), '%b %y'),
                    ' - ',
                    DATE_FORMAT(MAX(report_month), '%b %y')
                  ) AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY YEAR(report_month), QUARTER(report_month)
                ORDER BY MAX(report_month) DESC
            """)
        else:
            # SQLite
             sql = text("""
                SELECT
                  strftime('%m-%Y', MIN(report_month)) || ' - ' || strftime('%m-%Y', MAX(report_month)) AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY strftime('%Y', report_month), (CAST(strftime('%m', report_month) AS INTEGER) + 2) / 3
                ORDER BY MAX(report_month) DESC
            """)

    elif type == "H":
        if DB_TYPE == "mysql":
            sql = text("""
                SELECT
                  MAX(CASE
                    WHEN MONTH(report_month) BETWEEN 4 AND 9
                      THEN CONCAT('Apr ', YEAR(report_month), ' - Sep ', YEAR(report_month))
                    ELSE
                      CONCAT('Oct ', YEAR(report_month)-1, ' - Mar ', YEAR(report_month))
                  END) AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY
                  CASE
                    WHEN MONTH(report_month) BETWEEN 4 AND 9
                      THEN CONCAT('H1-', YEAR(report_month))
                    ELSE
                      CONCAT('H2-', YEAR(report_month))
                  END
                ORDER BY MAX(report_month) DESC
            """)
        else:
            # SQLite
            sql = text("""
                SELECT
                  CASE
                    WHEN CAST(strftime('%m', report_month) AS INTEGER) BETWEEN 4 AND 9
                      THEN 'Apr ' || strftime('%Y', report_month) || ' - Sep ' || strftime('%Y', report_month)
                    ELSE
                      'Oct ' || CAST(strftime('%Y', report_month, '-1 year') AS TEXT) || ' - Mar ' || strftime('%Y', report_month)
                  END AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY
                  CASE
                    WHEN CAST(strftime('%m', report_month) AS INTEGER) BETWEEN 4 AND 9
                      THEN 'H1-' || strftime('%Y', report_month)
                    ELSE
                      'H2-' || strftime('%Y', report_month)
                  END
                ORDER BY MAX(report_month) DESC
            """)

    elif type == "FY":
        if DB_TYPE == "mysql":
            sql = text("""
                SELECT
                  CONCAT(
                    'Apr ', MAX(YEAR(report_month))-1,
                    ' - Mar ', MAX(YEAR(report_month))
                  ) AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY YEAR(report_month)
                ORDER BY MAX(report_month) DESC
            """)
        else:
            # SQLite
            sql = text("""
                SELECT
                  'Apr ' || CAST(strftime('%Y', report_month, '-1 year') AS TEXT) || ' - Mar ' || strftime('%Y', report_month) AS label,
                  MIN(report_month) AS start_date,
                  MAX(report_month) AS end_date
                FROM irdai_monthly_data
                GROUP BY strftime('%Y', report_month)
                ORDER BY MAX(report_month) DESC
            """)

    else:
        raise HTTPException(status_code=400, detail="Invalid period type")

    return db.execute(sql).mappings().all()


# ======================================================
# 3️⃣ DASHBOARD TOTALS (PUBLIC / PRIVATE / GRAND)
# ======================================================
@router.get("/dashboard/totals")
def get_dashboard_totals(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          -- FYP
          SUM(CASE WHEN insurer_name = 'LIC of India' THEN fyp_current ELSE 0 END)     AS fyp_public,
          SUM(CASE WHEN insurer_name = 'Private Total' THEN fyp_current ELSE 0 END)   AS fyp_private,
          SUM(CASE WHEN insurer_name = 'Grand Total' THEN fyp_current ELSE 0 END)     AS fyp_grand,

          -- SA
          SUM(CASE WHEN insurer_name = 'LIC of India' THEN sa_current ELSE 0 END)     AS sa_public,
          SUM(CASE WHEN insurer_name = 'Private Total' THEN sa_current ELSE 0 END)   AS sa_private,
          SUM(CASE WHEN insurer_name = 'Grand Total' THEN sa_current ELSE 0 END)     AS sa_grand,

          -- POLICIES
          SUM(CASE WHEN insurer_name = 'LIC of India' THEN pol_current ELSE 0 END)    AS nop_public,
          SUM(CASE WHEN insurer_name = 'Private Total' THEN pol_current ELSE 0 END)  AS nop_private,
          SUM(CASE WHEN insurer_name = 'Grand Total' THEN pol_current ELSE 0 END)    AS nop_grand,

          -- LIVES
          SUM(CASE WHEN insurer_name = 'LIC of India' THEN lives_current ELSE 0 END)  AS nol_public,
          SUM(CASE WHEN insurer_name = 'Private Total' THEN lives_current ELSE 0 END)AS nol_private,
          SUM(CASE WHEN insurer_name = 'Grand Total' THEN lives_current ELSE 0 END)  AS nol_grand

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND category = insurer_name
    """)

    r = db.execute(sql, {
        "start_date": start_date,
        "end_date": end_date
    }).mappings().first()

    return {
        "FYP": {
            "public": r["fyp_public"] or 0,
            "private": r["fyp_private"] or 0,
            "grand": r["fyp_grand"] or 0,
        },
        "SA": {
            "public": r["sa_public"] or 0,
            "private": r["sa_private"] or 0,
            "grand": r["sa_grand"] or 0,
        },
        "NOP": {
            "public": r["nop_public"] or 0,
            "private": r["nop_private"] or 0,
            "grand": r["nop_grand"] or 0,
        },
        "NOL": {
            "public": r["nol_public"] or 0,
            "private": r["nol_private"] or 0,
            "grand": r["nol_grand"] or 0,
        }
    }


# ======================================================
# 4️⃣ PREMIUM TYPE SUMMARY (PUBLIC vs PRIVATE)
# ======================================================
@router.get("/dashboard/premium-type-summary")
def get_premium_type_summary(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          CASE
            WHEN insurer_name = 'LIC of India' THEN 'Public'
            ELSE 'Private'
          END AS insurer_type,

          category AS premium_type,

          SUM(fyp_current)   AS fyp,
          SUM(sa_current)    AS sum_assured,
          SUM(lives_current) AS lives,
          SUM(pol_current)   AS policies

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )

        GROUP BY insurer_type, category
        ORDER BY insurer_type, category
    """)

    return db.execute(sql, {
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# ======================================================
# 5️⃣ METRIC-WISE PREMIUM BREAKUP
# ======================================================

@router.get("/dashboard/metric-wise-premium")
def get_metric_wise_premium_breakup(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT 'FYP' AS metric, category AS premium_type, SUM(fyp_current) AS value
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        SELECT 'SA', category, SUM(sa_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        SELECT 'NOP', category, SUM(pol_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        SELECT 'NOL', category, SUM(lives_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        ORDER BY metric, premium_type
    """)

    return db.execute(sql, {
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# ======================================================
# COMPANYWISE Page
# 6️⃣ INSURER LIST
@router.get("/company/insurers")
def get_company_list(db: Session = Depends(get_db)):
    sql = text("""
        SELECT DISTINCT insurer_name
        FROM irdai_monthly_data
        WHERE insurer_name NOT IN ('Private Total', 'Grand Total')
        ORDER BY insurer_name
    """)
    rows = db.execute(sql).fetchall()
    return [{"label": r[0], "value": r[0]} for r in rows]


# 7️⃣ COMPANY TOTALS
@router.get("/company/totals")
def get_company_totals(
    insurer_name: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          SUM(fyp_current)   AS fyp,
          SUM(sa_current)    AS sa,
          SUM(pol_current)   AS nop,
          SUM(lives_current) AS nol
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category = insurer_name
    """)

    row = db.execute(sql, {
        "insurer_name": insurer_name,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().first()
    
    if not row:
        return {}
        
    # Normalize keys to lowercase to ensure consistency
    return {k.lower(): v for k, v in row.items()}




# 8️⃣ COMPANY PREMIUM TYPE BREAKUP
@router.get("/company/premium-type")
def get_company_premium_type_breakup(
    insurer_name: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          category AS premium_type,
          SUM(fyp_current)   AS fyp,
          SUM(sa_current)    AS sa,
          SUM(pol_current)   AS nop,
          SUM(lives_current) AS nol
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category <> insurer_name
        GROUP BY category
        ORDER BY category
    """)

    return db.execute(sql, {
        "insurer_name": insurer_name,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# 9️⃣ COMPANY METRIC-WISE PREMIUM BREAKUP
@router.get("/company/metric-wise-premium")
def get_company_metric_wise_premium(
    insurer_name: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        /* ======================
           FYP
        ====================== */
        SELECT
          'FYP' AS metric,
          category AS premium_type,
          SUM(fyp_current) AS value
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        /* ======================
           SA
        ====================== */
        SELECT
          'SA',
          category,
          SUM(sa_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        /* ======================
           NOP
        ====================== */
        SELECT
          'NOP',
          category,
          SUM(pol_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        UNION ALL

        /* ======================
           NOL
        ====================== */
        SELECT
          'NOL',
          category,
          SUM(lives_current)
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY category

        ORDER BY metric, premium_type
    """)

    return db.execute(sql, {
        "insurer_name": insurer_name,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# 1️⃣0️⃣ PREMIUM TYPES


@router.get("/premium/types")
def get_premium_types():
    return [
        {"label": "Individual Single Premium",
            "value": "Individual Single Premium"},
        {"label": "Individual Non-Single Premium",
            "value": "Individual Non-Single Premium"},
        {"label": "Group Single Premium", "value": "Group Single Premium"},
        {"label": "Group Non-Single Premium", "value": "Group Non-Single Premium"},
        {"label": "Group Yearly Renewable Premium",
            "value": "Group Yearly Renewable Premium"},
    ]


# 1️⃣1️⃣ PREMIUM WISE COMPANIES
@router.get("/premium/companies")
def get_premium_wise_companies(
    premium_type: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          insurer_name,
          SUM(fyp_current)   AS fyp,
          SUM(sa_current)    AS sa,
          SUM(pol_current)   AS nop,
          SUM(lives_current) AS nol
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND category = :premium_type
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
        GROUP BY insurer_name
        ORDER BY fyp DESC
    """)

    return db.execute(sql, {
        "premium_type": premium_type,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# 1️⃣2️⃣ PREMIUM GRAND TOTALS
@router.get("/premium/grand-totals")
def get_premium_grand_totals(
    premium_type: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          SUM(fyp_current)   AS fyp,
          SUM(sa_current)    AS sa,
          SUM(pol_current)   AS nop,
          SUM(lives_current) AS nol
        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND category = :premium_type
          AND insurer_name NOT IN ('Private Total', 'Grand Total')
    """)

    r = db.execute(sql, {
        "premium_type": premium_type,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().first()

    return {
        "premium_type": premium_type,
        "FYP": r["fyp"] or 0,
        "SA": r["sa"] or 0,
        "NOP": r["nop"] or 0,
        "NOL": r["nol"] or 0
    }


# 1️⃣3️⃣ COMPANY PREMIUM MARKET SHARE
@router.get("/market-share/company-premium")
def get_company_premium_market_share(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        /* =========================
           COMPANY TOTAL ROWS
        ========================= */
        SELECT
          insurer_name,
          NULL AS premium_type,

          SUM(fyp_market_share)   AS fyp_pct,
          SUM(pol_market_share)   AS nop_pct,
          SUM(lives_market_share) AS nol_pct,
          SUM(sa_market_share)    AS sa_pct,

          0 AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND category = insurer_name
          AND insurer_name NOT IN ('Private Total', 'Grand Total')

        GROUP BY insurer_name

        UNION ALL

        /* =========================
           PREMIUM BREAKUP ROWS
        ========================= */
        SELECT
          insurer_name,
          category AS premium_type,

          SUM(fyp_market_share)   AS fyp_pct,
          SUM(pol_market_share)   AS nop_pct,
          SUM(lives_market_share) AS nol_pct,
          SUM(sa_market_share)    AS sa_pct,

          1 AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )

        GROUP BY insurer_name, category
        ORDER BY insurer_name, row_order, premium_type
    """)

    return db.execute(sql, {
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# 1️⃣4️⃣ PREMIUM MARKET SHARE BY INSURER
@router.get("/market-share/premium-by-insurer")
def get_premium_market_share_by_insurer(
    insurer_name: str,
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        /* =========================
           TOTAL ROW (REAL TOTAL)
        ========================= */
        SELECT
          insurer_name,
          category AS premium_type,

          SUM(fyp_market_share)   AS fyp_pct,
          SUM(pol_market_share)   AS nop_pct,
          SUM(lives_market_share) AS nol_pct,
          SUM(sa_market_share)    AS sa_pct,

          0 AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category NOT IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )
        GROUP BY insurer_name, category
        HAVING COUNT(*) > 0

        UNION ALL

        /* =========================
           PREMIUM TYPE ROWS
        ========================= */
        SELECT
          insurer_name,
          category AS premium_type,

          SUM(fyp_market_share)   AS fyp_pct,
          SUM(pol_market_share)   AS nop_pct,
          SUM(lives_market_share) AS nol_pct,
          SUM(sa_market_share)    AS sa_pct,

          1 AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND category IN (
            'Individual Single Premium',
            'Individual Non-Single Premium',
            'Group Single Premium',
            'Group Non-Single Premium',
            'Group Yearly Renewable Premium'
          )

        GROUP BY insurer_name, category
        ORDER BY row_order, premium_type
    """)

    return db.execute(sql, {
        "insurer_name": insurer_name,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()


# 1️⃣4️⃣ GROWTH METRIC TYPES
@router.get("/growth/metric-types")
def get_growth_metric_types():
    return [
        {
            "label": "First Year Premium",
            "value": "FYP"
        },
        {
            "label": "Sum Assured",
            "value": "SA"
        },
        {
            "label": "No. of Policies",
            "value": "NOP"
        },
        {
            "label": "No. of Lives",
            "value": "NOL"
        }
    ]

# 1️⃣5️⃣ COMPANY PREMIUM GROWTH

# 1️⃣5️⃣ COMPANY PREMIUM GROWTH


@router.get("/growth/company-premium")
def get_company_premium_growth(
    insurer_name: str,
    metric: str,   # FYP | SA | NOP | NOL
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    metric_map = {
        "FYP": {
            "prev": "fyp_prev",
            "cur": "fyp_current",
            "growth": "fyp_growth",
            "ytd_prev": "fyp_ytd_prev",
            "ytd": "fyp_ytd_current",
            "ytd_growth": "fyp_growth_ytd",
            "market_share": "fyp_market_share",
        },
        "SA": {
            "prev": "sa_prev",
            "cur": "sa_current",
            "growth": "sa_growth",
            "ytd_prev": "sa_ytd_prev",
            "ytd": "sa_ytd_current",
            "ytd_growth": "sa_growth_ytd",
            "market_share": "sa_market_share",
        },
        "NOP": {
            "prev": "pol_prev",
            "cur": "pol_current",
            "growth": "pol_growth",
            "ytd_prev": "pol_ytd_prev",
            "ytd": "pol_ytd_current",
            "ytd_growth": "pol_growth_ytd",
            "market_share": "pol_market_share",
        },
        "NOL": {
            "prev": "lives_prev",
            "cur": "lives_current",
            "growth": "lives_growth",
            "ytd_prev": "lives_ytd_prev",
            "ytd": "lives_ytd_current",
            "ytd_growth": "lives_growth_ytd",
            "market_share": "lives_market_share",
        },
    }

    if metric not in metric_map:
        raise HTTPException(status_code=400, detail="Invalid metric")

    m = metric_map[metric]

    sql = text(f"""
        SELECT
          insurer_name,
          category AS premium_type,

          SUM({m["prev"]})         AS previous_value,
          SUM({m["cur"]})          AS current_value,
          SUM({m["growth"]})       AS growth_pct,

          SUM({m["ytd_prev"]})     AS ytd_previous_value,
          SUM({m["ytd"]})          AS ytd_value,
          SUM({m["ytd_growth"]})   AS ytd_growth_pct,

          SUM({m["market_share"]}) AS market_share,

          CASE
            WHEN category NOT IN (
              'Individual Single Premium',
              'Individual Non-Single Premium',
              'Group Single Premium',
              'Group Non-Single Premium',
              'Group Yearly Renewable Premium'
            ) THEN 0   -- TOTAL row
            ELSE 1
          END AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name = :insurer_name
          AND (
            category NOT IN (
              'Individual Single Premium',
              'Individual Non-Single Premium',
              'Group Single Premium',
              'Group Non-Single Premium',
              'Group Yearly Renewable Premium'
            )
            OR category IN (
              'Individual Single Premium',
              'Individual Non-Single Premium',
              'Group Single Premium',
              'Group Non-Single Premium',
              'Group Yearly Renewable Premium'
            )
          )

        GROUP BY insurer_name, category
        ORDER BY row_order, premium_type
    """)

    return db.execute(sql, {
        "insurer_name": insurer_name,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()

# 1️⃣6️⃣ MONTHWISE – ALL COMPANIES – ALL METRICS
@router.get("/monthwise/all-companies-all-metrics")
def get_monthwise_all_companies_all_metrics(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db)
):
    sql = text("""
        SELECT
          insurer_name,
          category AS premium_type,

          /* ========== FYP ========== */
          fyp_prev          AS fyp_previous,
          fyp_current       AS fyp_current,
          fyp_growth        AS fyp_growth,
          fyp_ytd_prev      AS fyp_ytd_previous,
          fyp_ytd_current   AS fyp_ytd_current,
          fyp_growth_ytd    AS fyp_ytd_growth,
          fyp_market_share  AS fyp_market_share,

          /* ========== SA ========== */
          sa_prev           AS sa_previous,
          sa_current        AS sa_current,
          sa_growth         AS sa_growth,
          sa_ytd_prev       AS sa_ytd_previous,
          sa_ytd_current    AS sa_ytd_current,
          sa_growth_ytd     AS sa_ytd_growth,
          sa_market_share   AS sa_market_share,

          /* ========== NOP ========== */
          pol_prev          AS nop_previous,
          pol_current       AS nop_current,
          pol_growth        AS nop_growth,
          pol_ytd_prev      AS nop_ytd_previous,
          pol_ytd_current   AS nop_ytd_current,
          pol_growth_ytd    AS nop_ytd_growth,
          pol_market_share  AS nop_market_share,

          /* ========== NOL ========== */
          lives_prev        AS nol_previous,
          lives_current     AS nol_current,
          lives_growth      AS nol_growth,
          lives_ytd_prev    AS nol_ytd_previous,
          lives_ytd_current AS nol_ytd_current,
          lives_growth_ytd  AS nol_ytd_growth,
          lives_market_share AS nol_market_share

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date

        ORDER BY
          insurer_name,
          CASE
            WHEN category NOT IN (
              'Individual Single Premium',
              'Individual Non-Single Premium',
              'Group Single Premium',
              'Group Non-Single Premium',
              'Group Yearly Renewable Premium'
            ) THEN 0   -- TOTAL rows first
            ELSE 1
          END,
          category
    """)

    return db.execute(sql, {
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()
# 1️⃣7️⃣ PRIVATE vs PUBLIC TABLE
@router.get("/pvt-vs-public/table")
def get_pvt_vs_public_table(
    start_date: str,
    end_date: str,
    sector: str = "BOTH",        # PRIVATE | PUBLIC | BOTH
    premium_type: str = "ALL",   # ALL | Individual Single Premium | ...
    db: Session = Depends(get_db)
):
    # ----------------------------
    # Sector filter
    # ----------------------------
    if sector == "PRIVATE":
        insurer_filter = "('Private Total')"
    elif sector == "PUBLIC":
        insurer_filter = "('LIC of India')"
    else:  # BOTH
        insurer_filter = "('Grand Total','Private Total','LIC of India')"

    # ----------------------------
    # Premium type filter
    # ----------------------------
    if premium_type == "ALL":
        premium_filter = """
            category = insurer_name
            OR category IN (
              'Individual Single Premium',
              'Individual Non-Single Premium',
              'Group Single Premium',
              'Group Non-Single Premium',
              'Group Yearly Renewable Premium'
            )
        """
    else:
        premium_filter = "category = :premium_type"

    sql = text(f"""
        SELECT
          insurer_name,
          category AS row_name,

          SUM(fyp_current)   AS fyp,
          SUM(pol_current)   AS nop,
          SUM(lives_current) AS nol,
          SUM(sa_current)    AS sa,

          CASE
            WHEN insurer_name = 'Grand Total' THEN 0
            WHEN insurer_name = 'Private Total' THEN 1
            WHEN insurer_name = 'LIC of India' THEN 2
            ELSE 9
          END AS section_order,

          CASE
            WHEN category = insurer_name THEN 0
            ELSE 1
          END AS row_order

        FROM irdai_monthly_data
        WHERE report_month BETWEEN :start_date AND :end_date
          AND insurer_name IN {insurer_filter}
          AND ({premium_filter})

        GROUP BY insurer_name, category

        ORDER BY
          section_order,
          row_order,
          category
    """)

    params = {
        "start_date": start_date,
        "end_date": end_date,
    }

    if premium_type != "ALL":
        params["premium_type"] = premium_type

    return db.execute(sql, params).mappings().all()

  # 1️⃣8️⃣ PEER INSURERS LIST

@router.get("/dropdown/insurers")
def get_insurer_dropdown(db: Session = Depends(get_db)):
    sql = text("""
        SELECT DISTINCT insurer_name
        FROM irdai_monthly_data
        WHERE insurer_name NOT IN ('Grand Total', 'Private Total')
        AND (insurer_name LIKE '%Limited%' OR insurer_name LIKE 'LIC%') 
        ORDER BY insurer_name
    """)
    insurers = [r[0] for r in db.execute(sql).all()]

    return {
        "max_select": 5,
        "options": insurers
    }


@router.get("/peers/comparison")
def get_peer_comparison(
    insurers: list[str] = Query(..., description="Select up to 5 insurers"),
    metric: str = Query(..., description="FYP | SA | NOP | NOL"),
    premium_type: str = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...),
    db: Session = Depends(get_db)
):
    # ----------------------------
    # VALIDATIONS
    # ----------------------------
    if len(insurers) > 5:
        raise HTTPException(
            status_code=400,
            detail="You can compare a maximum of 5 insurers"
        )

    metric_map = {
        "FYP": "fyp_current",
        "SA": "sa_current",
        "NOP": "pol_current",
        "NOL": "lives_current",
    }

    if metric not in metric_map:
        raise HTTPException(status_code=400, detail="Invalid metric")

    allowed_premium_types = [
        "Individual Single Premium",
        "Individual Non-Single Premium",
        "Group Single Premium",
        "Group Non-Single Premium",
        "Group Yearly Renewable Premium",
    ]

    if premium_type not in allowed_premium_types:
        raise HTTPException(status_code=400, detail="Invalid premium type")

    metric_column = metric_map[metric]

    # ----------------------------
    # SQL (SAFE IN CLAUSE)
    # ----------------------------
    sql = (
        text(f"""
            SELECT
              insurer_name,
              SUM({metric_column}) AS value
            FROM irdai_monthly_data
            WHERE report_month BETWEEN :start_date AND :end_date
              AND insurer_name IN :insurers
              AND category = :premium_type
            GROUP BY insurer_name
            ORDER BY insurer_name
        """)
        .bindparams(bindparam("insurers", expanding=True))
    )

    rows = db.execute(sql, {
        "insurers": insurers,
        "premium_type": premium_type,
        "start_date": start_date,
        "end_date": end_date
    }).mappings().all()

    # ----------------------------
    # RESPONSE
    # ----------------------------
    return {
        "metric": metric,
        "premium_type": premium_type,
        "period": {
            "start_date": start_date,
            "end_date": end_date
        },
        "insurers": insurers,
        "data": {r["insurer_name"]: r["value"] for r in rows}
    }